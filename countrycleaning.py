import codecs, json
import pandas as pd
from bs4 import BeautifulSoup
import openpyxl
import time, pickle
from others import create_excel_file, print_df_to_excel
import random
import requests
from fake_useragent import UserAgent
from stem import Signal
from stem.control import Controller
from fuzzywuzzy import fuzz
from fuzzywuzzy import process
import re
import string

start = time.time()

# First index is 1, index is number of current authors
indextostart = 1

data_store_columns = ['country1', 'country2', 'country3', 'country4', 'country5', 'country6', 'country7', 'country8'
                      , 'country9', 'country10', 'country11', 'country12', 'country13']

data_store_columns2 = ['citystate1', 'citystate2', 'citystate3', 'citystate4', 'citystate5', 'citystate6', 'citystate7',
                      'citystate8', 'citystate9', 'citystate10', 'citystate11', 'citystate12', 'citystate13']

# Json from https://github.com/dr5hn/countries-states-cities-database

dfcountries = pd.read_json('countries.json'.format(1), encoding='utf-8')
dfstates = pd.read_json('states.json'.format(1), encoding='utf-8')
dfcities = pd.read_json('cities.json'.format(1), encoding='utf-8')

countrylist = dfcountries['name']
statelist = dfstates['name']
#print(statelist)
citylist = dfcities['name']
combinedcitystatelist = statelist.append(citylist)

def locationchecker(input):
    # print(dfcities.head())
    # print(dfcountries.loc[dfcountries['name'] == 'Spain', 'iso2'].iloc[0])
    # location dummy, 0 = country, 1 = state, 2 = city, 3 = cannot be determined
    locationdummy = 3
    try:
        country_code = dfcities.loc[dfcities['name'] == input, 'country_code'].iloc[0]
        locationdummy = 2
    except:
        pass
    try:
        country_code = dfstates.loc[dfstates['name'] == input, 'country_code'].iloc[0]
        locationdummy = 1
    except:
        pass
    try:
        country_code = dfcountries.loc[dfcountries['name'] == input, 'iso_2'].iloc[0]
        locationdummy = 0
    except:
        pass
    output = ''
    if locationdummy == 0:
        output = input + ' (country)'
    if locationdummy == 1:
        output = input + ' (state)'
    if locationdummy == 2:
        output = input + ' (city)'
    if locationdummy == 3:
        output = input + ' (cannot be determined)'
    #print(output)

    return locationdummy

with open('./pkl/IDEASaffiliationSecondVariant.pkl', 'rb') as handle:
    data_store = pickle.load(handle)
df = pd.DataFrame(data=data_store[1], columns=data_store[0])
locationdata = df['locationdata']
delimitedlocation = []
for i in range(len(locationdata)):
    textsplit = re.split('/ |/', locationdata[i])
    if '' in textsplit:
        textsplit.remove('')
    for i in range(len(textsplit)):
        textsplit[i] = textsplit[i].strip()
    delimitedlocation.append(textsplit)
#print(delimitedlocation)

#locationselected = 'INDIA VISAKHAPATNAM'
#print(process.extract(locationselected, countrylist, limit=1)[0])
#word = 'India'
#wordcap = word.upper()
#print(locationselected.replace(wordcap, ''))
#wordtry = 'demand side management of energy storage'
#print(string.capwords(wordtry))

countrydataoverall = []
citystatedataoverall = []
for authors in range(len(delimitedlocation)):
    countrydata = []
    citystatedata = []
    for location in range(len(delimitedlocation[authors])):
        #print(delimitedlocation[authors][location])
        countrytext = delimitedlocation[authors][location]
        if countrytext == 'NA':
            countrydata.append('NA')
            citystatedata.append('NA')
            break
        if '(United States)' in countrytext:
            country = 'United States'
            citystate = countrytext.replace('(United States)', '')
            citystate = citystate.replace(',', '')
        elif '(France)' in countrytext:
            country = 'France'
            citystate = countrytext.replace('(France)', '')
            citystate = citystate.replace(',', '')
        else:
            country1, score, index = process.extract(delimitedlocation[authors][location], countrylist, limit=1)[0]
            if country1 in delimitedlocation[authors][location]:
                country = country1
            else:
                country = 'No Country'
            citystate = countrytext.replace(country1, '')
            countrycap = country1.upper()
            citystate = citystate.replace(countrycap, '')
            citystate = citystate.replace(',', '')
            citystate = citystate.replace('-', '')
            citystate = citystate.lower()
            citystate = citystate.title()
        #data = process.extract(delimitedlocation[authors][location], combinedcitystatelist, limit=5)
        #citystates = []
        #for i in range(5):
        #    citystates.append(data[i][0])
        #citystate = 'No State/City'
        #for i in range(5):
        #    if citystates[i] in delimitedlocation[authors][location]:
        #        citystate = citystates[i]
        #print(country)
        #print(citystate)
        countrydata.append(country)
        citystatedata.append(citystate)
    while len(countrydata) < 13:
        countrydata.append('')
    while len(citystatedata) < 13:
        citystatedata.append('')
    countrydataoverall.append(countrydata)
    citystatedataoverall.append(citystatedata)

    with open('IDEAScountrydata{}.pkl'.format(indextostart), 'wb') as handle:
        pickle.dump([data_store_columns, countrydataoverall], handle, protocol=pickle.HIGHEST_PROTOCOL)

    with open('IDEAScitystatedata{}.pkl'.format(indextostart), 'wb') as handle:
        pickle.dump([data_store_columns2, citystatedataoverall], handle, protocol=pickle.HIGHEST_PROTOCOL)

    print('Progress: {} out of {} done'.format(authors + indextostart, len(delimitedlocation)))

    #if authors > 2:
    #    break


write_excel = create_excel_file('./results/{}_results.xlsx'.format('IDEAScountrydata'))
wb = openpyxl.load_workbook(write_excel)
ws = wb[wb.sheetnames[-1]]
print_df_to_excel(df=pd.DataFrame(data=countrydataoverall, columns=data_store_columns), ws=ws)
wb.save(write_excel)


write_excel = create_excel_file('./results/{}_results.xlsx'.format('IDEAScitystatedata'))
wb = openpyxl.load_workbook(write_excel)
ws = wb[wb.sheetnames[-1]]
print_df_to_excel(df=pd.DataFrame(data=citystatedataoverall, columns=data_store_columns2), ws=ws)
wb.save(write_excel)


elapsed = (time.time() - start) / 3600
print(f"Elapsed time: {elapsed} hours")